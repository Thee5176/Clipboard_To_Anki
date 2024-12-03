import time

import pyperclip
import pykakasi

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from deep_translator import GoogleTranslator

"""
/ Step 1: parse clipboard > py_list
/ Step 2: loop until KeyboardInterupt
TODO Step 3: select sheet_name
Step 4: iterate through py_list
    / 4.1; parse py_list's item > column A
    / 4.2; find furigana > column B
    / 4.3; translate to english > append column B
    / 4.4; tag > column C
"""

#Step 4
def append_to_spreadsheet(filename, column_no, dataset):
    try:
        workbook = load_workbook(filename)
        sheet = workbook.active
    except FileNotFoundError:
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
    
        #Step 4.2
    if type(dataset) == str:
        #find last row of prev column
        column = get_column_letter(column_no)
        row_with_data = [cell for cell in sheet[column] if cell.value is not None]
        next_row = int(row_with_data[-1].row) + 1  if row_with_data else 1

        prev_column = get_column_letter(column_no - 1)
        prev_row_with_data = [cell for cell in sheet[prev_column] if cell.value is not None]   
        last_row = int(prev_row_with_data[-1].row) + 1

        for i in range(next_row, last_row):
            sheet[f"{column}{i}"] = dataset
        
    #Step 4.1, 4.3
    else:
        #find last row of that column
        column = get_column_letter(column_no)
        row_with_data = [cell for cell in sheet[column] if cell.value is not None]
        next_row = int(row_with_data[-1].row) + 1  if row_with_data else 1
        
        #append data from list
        for data in dataset:
            sheet[f"{column}{next_row}"] = data
            next_row = next_row + 1

    workbook.save(filename)

def find_furigana(dataset):
    furigana_list = []
    kks = pykakasi.kakasi()
    
    for data in dataset:
        result = kks.convert(data)       #result in dictionary of generator object
        furigana = "".join(item['hira'] for item in result) #convert generator obj > str
        furigana_list.append(furigana)
        
    return furigana_list

def find_translation(dataset):
    translation_list = []
    
    translator = GoogleTranslator(source='ja', target='en')
    translation_list = translator.translate_batch(dataset)
    
    return translation_list
    

def append_clipboard(filename):
    saved_words = []
    last_text = ""
    print("Clipboard monitor started. Press Ctrl+C to stop.")

    first_clip = True
    
    try:
        while True:   #FIXME non stop repeating??
            current_text = pyperclip.paste()
            if current_text != last_text and current_text.strip():
                
                if first_clip:
                    #skip first clipboard value
                    first_clip = False
                    print(f"ignored first clipboard item: {current_text}")
                    continue

                last_text = current_text
                print(f"Detect new clipboard item: {current_text}.")
                
                # append new word to list
                saved_words.append(current_text)
                print(f"Saved new item: {current_text}.")            
                
            time.sleep(0.5)

    except KeyboardInterrupt:
        print("\nClipboard monitor stopped.")
        
        #Front
        append_to_spreadsheet(filename, 1, saved_words)
        print(f"Parsed {len(saved_words)} clipboard items to spreadsheet")
        
        #Back #TODO keep the coresponse row
        furigana_list = find_furigana(saved_words)
        translation_list = find_translation(saved_words)
        
        back_list = [f"{furigana}　{translation}" for furigana,translation in zip(furigana_list,translation_list)]
        
        append_to_spreadsheet(filename, 2, back_list)
        print("Appended furigana")
        
        #Tags
        tag = input('Add a tag/tags: *separate with comma*')
        append_to_spreadsheet(filename, 3, tag)
        print("Appended tag")
        
if __name__ == "__main__":
    file = "Anki.xlsx"
    append_clipboard(file)
